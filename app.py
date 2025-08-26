import os
import json
import logging
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import uuid

# Configure logging
logging.basicConfig(level=logging.DEBUG)

app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev_secret_key_change_in_production")
CORS(app)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/timetable/course')
def timetable_course():
    return render_template('timetable_course.html')

@app.route('/timetable/class')
def timetable_class():
    return render_template('timetable_class.html')

@app.route('/manage')
def manage():
    return render_template('manage.html')

@app.route('/history')
def history():
    return render_template('history.html')

@app.route('/export/pdf', methods=['POST'])
def export_pdf():
    try:
        data = request.json
        
        # Create temporary file for PDF
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Create PDF document
        doc = SimpleDocTemplate(temp_file.name, pagesize=landscape(A4))
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Title']
        
        # Title
        title = f"Timetable - {data.get('courseName', 'N/A')} - {data.get('className', 'N/A')}"
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Create table data
        timetable_grid = data.get('timetableGrid', [])
        if not timetable_grid:
            elements.append(Paragraph("No timetable data available", styles['Normal']))
        else:
            # Headers
            headers = ['Time'] + [f'Day {i+1}' for i in range(len(timetable_grid[0]) if timetable_grid else 0)]
            table_data = [headers]
            
            # Rows
            for i, row in enumerate(timetable_grid):
                row_data = [f'Period {i+1}']
                for cell in row:
                    if cell.get('isBreak', False):
                        row_data.append('BREAK')
                    elif cell.get('subject'):
                        cell_text = f"{cell['subject']}\n{cell.get('teacher', '')}"
                        row_data.append(cell_text)
                    else:
                        row_data.append('Free')
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        return send_file(temp_file.name, as_attachment=True, download_name='timetable.pdf')
    
    except Exception as e:
        logging.error(f"PDF export error: {str(e)}")
        return jsonify({'error': 'Failed to generate PDF'}), 500

@app.route('/export/excel', methods=['POST'])
def export_excel():
    try:
        data = request.json
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Title
        title = f"Timetable - {data.get('courseName', 'N/A')} - {data.get('className', 'N/A')}"
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Time'] + [f'Day {i+1}' for i in range(data.get('workingDays', 5))]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Timetable data
        timetable_grid = data.get('timetableGrid', [])
        for row_idx, row in enumerate(timetable_grid, 4):
            ws.cell(row=row_idx, column=1, value=f'Period {row_idx-3}')
            
            for col_idx, cell_data in enumerate(row, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell_data.get('isBreak', False):
                    cell.value = 'BREAK'
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif cell_data.get('subject'):
                    subjectCode = cell_data.get('subjectCode', '')
                    subject = cell_data.get('subject', '')
                    teacher = cell_data.get('teacher', '')
                    cell.value = f"{subjectCode}\n{subject}\n{teacher}"
                else:
                    cell.value = 'Free'
                
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return send_file(temp_file.name, as_attachment=True, download_name='timetable.xlsx')
    
    except Exception as e:
        logging.error(f"Excel export error: {str(e)}")
        return jsonify({'error': 'Failed to generate Excel file'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
